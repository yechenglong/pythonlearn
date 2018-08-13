import openpyxl,logging
logging.basicConfig(level=logging.DEBUG,format=' %(asctime)s - %(levelname)s- %(message)s')
wb = openpyxl.load_workbook('123.xlsx')
l = wb.sheetnames
logging.debug(l)
sheet = wb['Sheet4']
logging.debug(sheet)
logging.debug(sheet['A1'].value)
logging.debug(sheet.max_row)
logging.debug(sheet.max_column)
